# -*- coding: utf-8 -*-
"""
📊 QUEUE-WISE CALL DATA BIFURCATION — Streamlit port of the Apps Script version.

User uploads the raw call report (CSV / Excel). The app parses the "Call Flow"
column and builds:
  1. Queue Wise Detail  — one row per call: 1st Agent / Queue / Talk Time, 2nd Agent..., Transfer Chain
  2. Agent Legs         — one row per agent-leg (long format, pivot-friendly)
  3. Queue Summary      — legs handled, unique agents, total & avg talk time per queue
  4. Agent x Queue      — matrix of legs handled by each agent in each queue

Output: downloadable Excel workbook (4 sheets) and/or push directly to a
Google Sheet (via the service account in st.secrets).

Talk time logic (same as Apps Script):
  agent's talk time = agent answer time → next "Inboundqueue" event (transfer), or Hangup if none.
"""

import io
import re
from datetime import datetime

import pandas as pd
import streamlit as st

# ─────────────────────────────────────────────────────────────
# Parsing helpers (1:1 port of the Apps Script logic)
# ─────────────────────────────────────────────────────────────
TS = r"(\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2})"
RE_QUEUE = re.compile(r"Inboundqueue: (.+?)\(" + TS + r"\)")
RE_AGENT = re.compile(r"Agent: (.+?)\(Answer\)\(" + TS + r"\)")
RE_HANGUP = re.compile(r"Hangup: \(" + TS + r"\)")

ORDINALS = ["1st", "2nd", "3rd", "4th", "5th", "6th", "7th", "8th",
            "9th", "10th", "11th", "12th"]


def parse_ts(s: str):
    try:
        return datetime.strptime(s, "%Y-%m-%d %H:%M:%S")
    except (ValueError, TypeError):
        return None


def to_date(v):
    """Fallback parser for the 'Call End Time' column."""
    if isinstance(v, datetime):
        return v
    if isinstance(v, pd.Timestamp):
        return v.to_pydatetime()
    if isinstance(v, str) and v.strip():
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S", "%d/%m/%Y %H:%M:%S"):
            try:
                return datetime.strptime(v.strip(), fmt)
            except ValueError:
                continue
        try:
            d = pd.to_datetime(v)
            return d.to_pydatetime() if not pd.isna(d) else None
        except Exception:
            return None
    return None


def parse_call_flow(cf: str):
    """Parse the Call Flow string into ordered events: queue / agent / hangup."""
    events = []
    for m in RE_QUEUE.finditer(cf):
        events.append({"pos": m.start(), "type": "queue",
                       "name": m.group(1).strip(), "ts": parse_ts(m.group(2))})
    for m in RE_AGENT.finditer(cf):
        events.append({"pos": m.start(), "type": "agent",
                       "name": m.group(1).strip(), "ts": parse_ts(m.group(2))})
    for m in RE_HANGUP.finditer(cf):
        events.append({"pos": m.start(), "type": "hangup",
                       "name": "", "ts": parse_ts(m.group(1))})
    events.sort(key=lambda e: e["pos"])
    return events


def sec_to_hms(sec):
    """seconds → 'H:MM:SS' string (blank if unknown)."""
    if sec is None or (isinstance(sec, float) and pd.isna(sec)):
        return ""
    sec = int(round(sec))
    h, rem = divmod(sec, 3600)
    m, s = divmod(rem, 60)
    return f"{h}:{m:02d}:{s:02d}"


# ─────────────────────────────────────────────────────────────
# Core processing
# ─────────────────────────────────────────────────────────────
def process_report(df: pd.DataFrame, col: dict):
    """Returns (calls, legs, max_agents). Mirrors buildQueueWiseReport()."""
    calls, legs = [], []
    max_agents = 0

    has_end = col["endTime"] in df.columns
    has_dir = col["direction"] in df.columns
    has_rec = col["recording"] in df.columns

    for _, row in df.iterrows():
        call_id = row.get(col["callId"], "")
        if pd.isna(call_id) or str(call_id).strip() == "":
            continue

        cf = row.get(col["callFlow"], "")
        cf = "" if pd.isna(cf) else str(cf)
        events = parse_call_flow(cf)

        hangup = None
        for e in events:
            if e["type"] == "hangup":
                hangup = e["ts"]
        final_end = hangup or (to_date(row[col["endTime"]]) if has_end else None)

        direction = ""
        if has_dir and not pd.isna(row.get(col["direction"])):
            direction = str(row[col["direction"]])
        call_legs = []

        for i, ev in enumerate(events):
            if ev["type"] != "agent":
                continue
            # queue = nearest preceding Inboundqueue event
            queue = None
            for j in range(i - 1, -1, -1):
                if events[j]["type"] == "queue":
                    queue = events[j]["name"]
                    break
            if not queue:
                queue = "Direct / Outbound" if direction == "Outbound" else "No Queue"
            # leg ends at next queue event (transfer) or hangup/end
            leg_end = None
            for k in range(i + 1, len(events)):
                if events[k]["type"] == "queue":
                    leg_end = events[k]["ts"]
                    break
            if leg_end is None:
                leg_end = final_end
            talk_sec = None
            if leg_end and ev["ts"]:
                talk_sec = max(0.0, (leg_end - ev["ts"]).total_seconds())
            call_legs.append({"agent": ev["name"], "queue": queue, "talkSec": talk_sec})

        max_agents = max(max_agents, len(call_legs))
        chain = " -> ".join(l["agent"] for l in call_legs)

        calls.append({
            "callId": call_id,
            "time": row.get(col["time"], ""),
            "direction": direction,
            "customer": row.get(col["customer"], ""),
            "convDur": "" if pd.isna(row.get(col["convDur"], "")) else str(row.get(col["convDur"], "")),
            "recording": (row.get(col["recording"], "") if has_rec else ""),
            "legs": call_legs,
            "chain": chain,
        })
        for n, l in enumerate(call_legs):
            legs.append({
                "Call ID": call_id,
                "Time": row.get(col["time"], ""),
                "Customer Number": row.get(col["customer"], ""),
                "Leg #": n + 1,
                "Agent": l["agent"],
                "Queue": l["queue"],
                "Talk Time (sec)": l["talkSec"],
            })

    return calls, legs, max_agents


def build_dataframes(calls, legs, max_agents):
    """Build the 4 output DataFrames."""
    # ---- Sheet 1: Queue Wise Detail ----
    head1 = ["Call ID", "Time", "Direction", "Customer Number",
             "Conversation Duration", "Recording"]
    for a in range(max_agents):
        o = ORDINALS[a] if a < len(ORDINALS) else f"{a + 1}th"
        head1 += [f"{o} Agent", f"{o} Agent Queue", f"{o} Agent Talk Time"]
    head1.append("Transfer Chain")

    rows1 = []
    for c in calls:
        v = [c["callId"], c["time"], c["direction"], c["customer"],
             c["convDur"], c["recording"]]
        for a in range(max_agents):
            if a < len(c["legs"]):
                l = c["legs"][a]
                v += [l["agent"], l["queue"], sec_to_hms(l["talkSec"])]
            else:
                v += ["", "", ""]
        v.append(c["chain"])
        rows1.append(v)
    df_detail = pd.DataFrame(rows1, columns=head1)

    # ---- Sheet 2: Agent Legs ----
    leg_cols = ["Call ID", "Time", "Customer Number", "Leg #",
                "Agent", "Queue", "Talk Time", "Talk Time (sec)"]
    if legs:
        df_legs = pd.DataFrame(legs)
        df_legs["Talk Time"] = df_legs["Talk Time (sec)"].apply(sec_to_hms)
        df_legs = df_legs[leg_cols]
    else:
        df_legs = pd.DataFrame(columns=leg_cols)

    # ---- Sheet 3: Queue Summary ----
    if not df_legs.empty:
        grp = df_legs.groupby("Queue", dropna=False)
        summary = grp.agg(**{
            "Call Legs Handled": ("Agent", "size"),
            "Unique Agents": ("Agent", "nunique"),
            "_total_sec": ("Talk Time (sec)", "sum"),
        }).reset_index().sort_values("Queue")
        summary["Total Talk Time"] = summary["_total_sec"].apply(sec_to_hms)
        summary["Avg Talk Time per Leg"] = summary.apply(
            lambda r: sec_to_hms(r["_total_sec"] / r["Call Legs Handled"])
            if r["Call Legs Handled"] else "0:00:00", axis=1)
        total_row = pd.DataFrame([{
            "Queue": "TOTAL",
            "Call Legs Handled": int(summary["Call Legs Handled"].sum()),
            "Unique Agents": "",
            "_total_sec": summary["_total_sec"].sum(),
            "Total Talk Time": sec_to_hms(summary["_total_sec"].sum()),
            "Avg Talk Time per Leg": "",
        }])
        df_summary = pd.concat([summary, total_row], ignore_index=True)[
            ["Queue", "Call Legs Handled", "Unique Agents",
             "Total Talk Time", "Avg Talk Time per Leg"]]
    else:
        df_summary = pd.DataFrame(columns=["Queue", "Call Legs Handled", "Unique Agents",
                                           "Total Talk Time", "Avg Talk Time per Leg"])

    # ---- Sheet 4: Agent x Queue matrix ----
    if not df_legs.empty:
        pivot = pd.pivot_table(df_legs, index="Agent", columns="Queue",
                               values="Call ID", aggfunc="size", fill_value=0)
        pivot = pivot.sort_index().sort_index(axis=1)
        pivot["Total Legs"] = pivot.sum(axis=1)
        talk = df_legs.groupby("Agent")["Talk Time (sec)"].sum()
        pivot["Total Talk Time"] = talk.reindex(pivot.index).apply(sec_to_hms)
        df_matrix = pivot.reset_index()
        df_matrix.columns.name = None
    else:
        df_matrix = pd.DataFrame(columns=["Agent", "Total Legs", "Total Talk Time"])

    return df_detail, df_legs, df_summary, df_matrix


def make_excel(df_detail, df_legs, df_summary, df_matrix) -> bytes:
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df_detail.to_excel(writer, sheet_name="Queue Wise Detail", index=False)
        df_legs.to_excel(writer, sheet_name="Agent Legs", index=False)
        df_summary.to_excel(writer, sheet_name="Queue Summary", index=False)
        df_matrix.to_excel(writer, sheet_name="Agent x Queue", index=False)
        # bold header + freeze top row on every sheet
        from openpyxl.styles import Font, PatternFill
        fill = PatternFill("solid", fgColor="1F4E79")
        font = Font(bold=True, color="FFFFFF")
        for ws in writer.book.worksheets:
            for cell in ws[1]:
                cell.font = font
                cell.fill = fill
            ws.freeze_panes = "A2"
    return buf.getvalue()


def push_to_gsheet(sheet_id: str, dfs: dict):
    """Overwrite/create one tab per output DataFrame in the destination sheet."""
    import gspread
    from google.oauth2.service_account import Credentials

    scopes = ["https://www.googleapis.com/auth/spreadsheets",
              "https://www.googleapis.com/auth/drive"]
    creds = Credentials.from_service_account_info(
        dict(st.secrets["gcp_service_account"]), scopes=scopes)
    gc = gspread.authorize(creds)
    ss = gc.open_by_key(sheet_id)

    for name, df in dfs.items():
        df_out = df.copy().astype(object).where(pd.notna(df), "")
        values = [[str(c) for c in df_out.columns]] + [
            [v if isinstance(v, (int, float)) else str(v) for v in row]
            for row in df_out.values.tolist()
        ]
        try:
            ws = ss.worksheet(name)
            ws.clear()
        except gspread.exceptions.WorksheetNotFound:
            ws = ss.add_worksheet(title=name, rows=max(len(values) + 5, 50),
                                  cols=max(len(values[0]) + 2, 10))
        ws.update(values, value_input_option="USER_ENTERED")
        ws.format("1:1", {
            "textFormat": {"bold": True,
                           "foregroundColor": {"red": 1, "green": 1, "blue": 1}},
            "backgroundColor": {"red": 0.12, "green": 0.31, "blue": 0.47},
        })
        ws.freeze(rows=1)
    return f"https://docs.google.com/spreadsheets/d/{sheet_id}"


# ─────────────────────────────────────────────────────────────
# UI
# ─────────────────────────────────────────────────────────────
st.title("📊 Queue Report Builder")
st.caption("Upload raw call report (CSV / Excel) → parses **Call Flow** → "
           "Queue Wise Detail • Agent Legs • Queue Summary • Agent × Queue")

with st.sidebar:
    st.header("⚙️ Column Mapping")
    st.caption("Header names as they appear in row 1 of your file")
    col = {
        "callId":    st.text_input("Call ID column", "Call ID"),
        "time":      st.text_input("Time column", "Time"),
        "direction": st.text_input("Direction column", "Direction"),
        "customer":  st.text_input("Customer Number column", "Customer Number"),
        "convDur":   st.text_input("Conversation Duration column", "Conversation Duration"),
        "recording": st.text_input("Recording column", "Recording"),
        "callFlow":  st.text_input("Call Flow column", "Call Flow"),
        "endTime":   st.text_input("Call End Time column", "Call End Time"),
    }

uploaded = st.file_uploader(
    "Upload raw call report",
    type=["csv", "xlsx", "xls"],
    help="The export that contains the Call Flow column",
)

if uploaded is None:
    st.info("👆 Upload your raw call report to begin. "
            "Required columns: **Call ID**, **Time**, **Customer Number**, "
            "**Conversation Duration**, **Call Flow** (others optional).")
    st.stop()

# ---- read file ----
try:
    if uploaded.name.lower().endswith(".csv"):
        df_raw = pd.read_csv(uploaded, dtype=str)
    else:
        df_raw = pd.read_excel(uploaded, dtype=str)
except Exception as e:
    st.error(f"File read nahi hui: {e}")
    st.stop()

df_raw.columns = [str(c).strip() for c in df_raw.columns]

# ---- validate required columns ----
required = ["callId", "time", "customer", "convDur", "callFlow"]
missing = [col[k] for k in required if col[k] not in df_raw.columns]
if missing:
    st.error("Yeh columns file mein nahi mile: **" + ", ".join(missing) +
             "**\n\nFile ke actual columns: " + ", ".join(df_raw.columns))
    st.stop()

st.success(f"✅ File loaded: **{uploaded.name}** — {len(df_raw)} rows")
with st.expander("👀 Raw data preview (first 5 rows)"):
    st.dataframe(df_raw.head(), use_container_width=True)

if st.button("🚀 Build Queue-Wise Report", type="primary"):
    with st.spinner("Call Flow parse ho raha hai..."):
        calls, legs, max_agents = process_report(df_raw, col)
        results = build_dataframes(calls, legs, max_agents)
    st.session_state["queue_results"] = results
    st.session_state["queue_stats"] = {
        "calls": len(calls), "legs": len(legs),
        "queues": (sorted(results[1]["Queue"].unique().tolist())
                   if not results[1].empty else []),
    }

if "queue_results" in st.session_state:
    df_detail, df_legs, df_summary, df_matrix = st.session_state["queue_results"]
    stats = st.session_state["queue_stats"]

    st.divider()
    c1, c2, c3 = st.columns(3)
    c1.metric("Calls processed", stats["calls"])
    c2.metric("Agent legs", stats["legs"])
    c3.metric("Queues found", len(stats["queues"]))
    if stats["queues"]:
        st.caption("Queues: " + " • ".join(stats["queues"]))

    t1, t2, t3, t4 = st.tabs(["Queue Wise Detail", "Agent Legs",
                              "Queue Summary", "Agent × Queue"])
    with t1:
        st.dataframe(df_detail, use_container_width=True)
    with t2:
        st.dataframe(df_legs, use_container_width=True)
    with t3:
        st.dataframe(df_summary, use_container_width=True)
    with t4:
        st.dataframe(df_matrix, use_container_width=True)

    st.divider()
    st.subheader("📤 Export")

    left, right = st.columns(2)

    with left:
        st.markdown("**Option A — Download Excel** (4 sheets)")
        xlsx_bytes = make_excel(df_detail, df_legs, df_summary, df_matrix)
        st.download_button(
            "⬇️ Download Queue Report.xlsx",
            data=xlsx_bytes,
            file_name="Queue Report.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    with right:
        st.markdown("**Option B — Push to Google Sheet**")
        if "gcp_service_account" not in st.secrets:
            st.info("Google Sheet push ke liye `[gcp_service_account]` secret add karo "
                    "(Streamlit Cloud → App → Settings → Secrets).")
        else:
            dest_id = st.text_input(
                "Destination Google Sheet ID",
                help="Sheet URL ka beech wala hissa: "
                     "docs.google.com/spreadsheets/d/<YEH_WALA>/edit — sheet ko service "
                     "account email ke saath Editor access se share karna zaroori hai.")
            if st.button("📤 Push 4 tabs to Google Sheet") and dest_id:
                try:
                    with st.spinner("Google Sheet mein likha ja raha hai..."):
                        url = push_to_gsheet(dest_id.strip(), {
                            "Queue Wise Detail": df_detail,
                            "Agent Legs": df_legs,
                            "Queue Summary": df_summary,
                            "Agent x Queue": df_matrix,
                        })
                    st.success(f"✅ Done! [Open Google Sheet]({url})")
                except Exception as e:
                    st.error(f"Push failed: {e}\n\n⚠️ Check karo ki destination sheet ko "
                             "service account email (secrets ke `client_email`) ke saath "
                             "**Editor** access diya hai.")
